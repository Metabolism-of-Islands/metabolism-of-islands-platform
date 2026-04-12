from core.models import *
from dataclasses import dataclass
from django.conf import settings
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.core.exceptions import ValidationError
from django.core.mail import send_mail
from django.core.validators import validate_email
from django.db.models import Count, Q, Subquery, OuterRef, CharField, Avg
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, get_object_or_404, redirect
from django.template.loader import render_to_string
from django.urls import reverse
from io import BytesIO
from itertools import combinations
from openpyxl.styles import Font
import csv
import openpyxl
import random
import re
import string

OPTAMOS_BG = [
    "pexels-altaf-shah-3143825-7751849.jpg",
    "pexels-ben-young-3465670-31953682.jpg",
    "pexels-ben-young-3465670-31953693.jpg",
    "pexels-kelly-2519392.jpg",
    "pexels-nacho-guillen-227263111-12041981.jpg",
    "pexels-polina-kovaleva-7648425.jpg",
    "pexels-simplyart4794-53798833-10900608.jpg",
    "pexels-stan-versluis-172635488-11034596.jpg",
    "pexels-tomfisk-1595108.jpg",
    "pexels-tomfisk-1605270.jpg",
    "pexels-tomfisk-2101137.jpg",
    "pexels-tomfisk-2246950.jpg",
    "pexels-tomfisk-3145153.jpg",
    "pexels-tomfisk-3256376.jpg",
    "pexels-tomfisk-3266775.jpg",
    "pexels-tomfisk-4198586.jpg",
    "pexels-tomfisk-5834061.jpg",
    "pexels-tomfisk-6926658.jpg",
    "pexels-tomfisk-7013401.jpg",
    "pexels-tomfisk-9940117.jpg",
    "pexels-vividcafe-681347.jpg",
]

# Random Index (RI) table (Saaty) used for consistency ratio calculations
RI_TABLE = {
    1: 0.00,
    2: 0.00,
    3: 0.58,
    4: 0.90,
    5: 1.12,
    6: 1.24,
    7: 1.32,
    8: 1.41,
    9: 1.45,
    10: 1.49,
}

@dataclass(frozen=True)
class ConsistencyResult:
    cr: float
    ci: float
    lambda_max: float

#### RANKING VALUE CONVERTER ####
# We need to calibrate the scores. 
# In AHP the default scale is 1-9, but we use -8 to 8 in our slider. That is done so we can use a native slider 
# where the 0 is the neutral value in between them both.
# But it means that a score of 0 in our system represents a score of 1 for both criteria in the AHP system
# And any other score (e.g. 6) represents a n+1 (e.g. 7 in the example) in our system
# A negative score simply means it's "in favor" of the other criteria, so we do the same procedure in reverse
# In the database we always store a single value (of alt1 vs alt2) which is 1-9 in regular AHP scoring (if alt2 
# is preferred, or -2 to -9 which simply means that alt1 is preferred. There will never a value of 0 being stored
# (or -1 for that matter) because the 1 is the score that means that neither is preferred.
def ranking_value_converter(value):
    if value < 0: 
        # This score is "in favor" of alternative 1, and we must subtract 1 to change the -1 to -8 range to 
        # -2 to -9
        return value-1
    else:
        # If the score is 0 or higher, we simply add 1, which means that the 0-8 range becomes 1-9
        return value+1 

def calculate_consistency_ratio(items_to_review, score_list):
    # Step 1: Load all items and scores
    items_to_review = items_to_review
    n = len(items_to_review)
    if n < 2:
        return ConsistencyResult(cr=0.0, ci=0.0, lambda_max=0.0)

    # Step 2: Build matrix
    matrix = np.ones((n, n))  # diagonal = 1 automatically
    crit_id_to_index = {c.id: idx for idx, c in enumerate(items_to_review)}

    for score in score_list:
        i = crit_id_to_index[score.id1]
        j = crit_id_to_index[score.id2]
        matrix[i, j] = score.value1
        matrix[j, i] = score.value2

    # Step 3: Compute principal eigenvalue
    eigenvalues, _ = np.linalg.eig(matrix)
    lambda_max = max(eigenvalues.real)

    # Step 4: Consistency Index (CI)
    ci = (lambda_max - n) / (n - 1)

    # Step 5: Random Index (RI)
    ri = RI_TABLE.get(n, 1.49) # Highly doubtful it goes beyond 10 but setting 1.49 if it does is an easy way to deal with that might it happen

    # Step 6: CR
    cr = ci / ri
    return ConsistencyResult(cr=cr, ci=ci, lambda_max=lambda_max)

def create_matrix(project, request):
    # Create a matrix with all criteria (values are 0 for each item)
    matrix_criteria = list(project.criteria.all())

    criteria_ids = [c.id for c in matrix_criteria]
    matrix = {
        c1.id: {c2.id: (1 if c1.id == c2.id else 0) for c2 in matrix_criteria} # Default if 0 if not set, but 1 for the diagonal values
        for c1 in matrix_criteria
    }

    # Load the actual scores into the matrix
    for score in OptamosCriteriaValue.objects.filter(criteria1__project=project, user=request.user):
        c1 = score.criteria1.id
        c2 = score.criteria2.id
        matrix[c1][c2] = score.value1
        matrix[c2][c1] = score.value2

    return matrix

def index(request):
    context = {
        "bg": random.choice(OPTAMOS_BG),
        "menu": "index",
    }
    return render(request, "optamos/index.html", context)

def about(request):
    context = {
        "bg": random.choice(OPTAMOS_BG),
        "menu": "about",
    }
    return render(request, "optamos/about.html", context)

def resources(request):
    context = {
        "bg": random.choice(OPTAMOS_BG),
        "menu": "resources",
    }
    return render(request, "optamos/resources.html", context)

def projects(request):
    if not request.user.is_authenticated:
        return redirect(f"{reverse("optamos:login")}?redirect={request.path}")

    access_qs = OptamosUser.objects.filter(project=OuterRef("pk"), user=request.user).values("level")[:1]
    projects = OptamosProject.objects_include_private.filter(users__user=request.user).annotate(
        level=Subquery(access_qs, output_field=CharField())
    )

    if "delete" in request.GET:
        project = projects.get(uid=request.GET["delete"], level="admin")
        project.delete()
        messages.success(request, f"Project {project.name} has been deleted.")
        return redirect(request.path)

    context = {
        "bg": random.choice(OPTAMOS_BG),
        "projects": projects,
        "menu": "projects",
    }
    return render(request, "optamos/projects.html", context)

def project_create(request):

    if not request.user.is_authenticated:
        return redirect(f"{reverse("optamos:login")}?redirect={request.path}")

    if request.method == "POST":

        csv_error = None

        # If the user uploads a csv file, then we will try and load the criteria/alternatives from this file
        # However, we must first check if it's a valid file. If it isn't, we remain on the same page and 
        # prompt the user to correct. If it's correct, then we save the project and redirect.
        if request.FILES.get("csv_file"):
            csv_file = request.FILES["csv_file"]
            if not csv_file.name.lower().endswith(".csv"):
                csv_error = "Your file was not a valid CSV file."
            else:
                try:
                    decoded_file = csv_file.read().decode("utf-8").splitlines()
                    reader = csv.DictReader(decoded_file)
                    required_headers = {"CRITERIA", "ALTERNATIVES"}
                    if not reader.fieldnames or not required_headers.issubset(set(reader.fieldnames)):
                        csv_error = "CSV must contain headers: CRITERIA, ALTERNATIVES. Your file was not valid and could not be loaded."
                    else:
                        valid_csv = True
                except Exception as e:
                    csv_error = f"Error processing file: {str(e)}"

        if csv_error:
            messages.error(request, csv_error)
            messages.warning(request,  "Please review the structure of <a href='/media/optamos/sample.csv'>our template</a> as a baseline.")
        else:
            project = OptamosProject()
            project.is_public = False
            project.name = request.POST.get("name")
            project.goal = request.POST.get("goal")
            project.description = request.POST.get("description")
            project.save()

            OptamosUser.objects.create(project=project, user=request.user, level="admin")

            if request.POST.getlist("alternative"):
                for each in request.POST.getlist("alternative"):
                    if each:
                        OptamosAlternative.objects.create(project=project, name=each)

            if request.POST.getlist("tag"):
                for each in request.POST.getlist("tag"):
                    if each:
                        OptamosTag.objects.create(project=project, name=each)

            previous_main_criteria = None
            position = 0
            if request.POST.getlist("criteria"):
                criteria = request.POST.getlist("criteria")
                descriptions = request.POST.getlist("desc_criteria")
                child = request.POST.getlist("child")
                for criterion, description, child in zip(criteria, descriptions, child):
                    if criterion:
                        if description == "":
                            description = None
                        position += 1
                        parent = None
                        if child == "true" and previous_main_criteria:
                            parent = previous_main_criteria
                        info = OptamosCriteria.objects.create(project=project, name=criterion, description=description, position=position, parent=parent)
                        if not info.parent:
                            previous_main_criteria = info

            if request.FILES.get("csv_file"):
                messages.success(request, "Your csv file was loaded.")
                position = 0
                previous_main_criteria = None
                for row in reader:
                    criteria = row.get("CRITERIA")
                    alternative = row.get("ALTERNATIVES")

                    if criteria:
                        position += 1
                        parent = None
                        if criteria.startswith("-") and previous_main_criteria:
                            parent = previous_main_criteria
                            criteria = criteria[1:]
                        info = OptamosCriteria.objects.create(project=project, name=criteria.strip(), position=position, parent=parent)
                        if not parent:
                            previous_main_criteria = info

                    if alternative:
                        OptamosAlternative.objects.create(project=project, name=alternative.strip())

                # If the user clicked the CSV upload button, then we will redirect the user to the
                # page where they can continue editing the project because they are not necessarily
                # done with editing.
                return redirect(reverse("optamos:project_settings", args=[project.uid]))
            else:
                messages.success(request, "Your project was created. Is this a group project? <a href='team/'>Invite your team</a>")
                return redirect(reverse("optamos:project", args=[project.uid]))

    context = {
        "bg": random.choice(OPTAMOS_BG),
        "menu": "projects",
    }
    return render(request, "optamos/project.create.html", context)

def project_settings(request, id):

    if not request.user.is_authenticated:
        return redirect(f"{reverse("optamos:login")}?redirect={request.path}")

    project = OptamosProject.objects_include_private.get(pk=id)
    user_access = OptamosUser.objects.filter(project=project, user=request.user, level="admin")
    if not user_access.exists():
        return redirect(reverse("optamos:access_denied") + "?role=admin&url=" + request.get_full_path())

    if request.method == "POST":
        project.name = request.POST.get("name")
        project.goal = request.POST.get("goal")
        project.description = request.POST.get("description")
        project.save()

        for each in project.alternatives.all():
            label = f"alternative_{each.id}"
            if request.POST.get(label):
                each.name = request.POST[label]
                each.save()
            else:
                each.delete()

        for each in project.tag_list.all():
            label = f"tag_{each.id}"
            if request.POST.get(label):
                each.name = request.POST[label]
                each.save()
            else:
                each.delete()

        position = 0
        for each in project.criteria.all():
            label = f"criteria_{each.id}"
            if request.POST.get(label):
                each.name = request.POST[label]
                each.description = request.POST.get(f"desc_criteria_{each.id}")
                if request.POST.get(f"child_{each.id}") == "true":
                    each.parent = previous_main_criteria
                else:
                    each.parent = None
                    previous_main_criteria = each
                position += 1
                each.position = position
                if each.description == "":
                    each.description = None
                each.save()
            else:
                each.delete()

        if request.POST.getlist("alternative"):
            for each in request.POST.getlist("alternative"):
                if each:
                    OptamosAlternative.objects.create(project=project, name=each)

        if request.POST.getlist("tag"):
            for each in request.POST.getlist("tag"):
                if each:
                    OptamosTag.objects.create(project=project, name=each)

        if request.POST.getlist("criteria"):
            criteria = request.POST.getlist("criteria")
            descriptions = request.POST.getlist("desc_criteria")
            child = request.POST.getlist("child")
            for criterion, description, child in zip(criteria, descriptions, child):
                if criterion:
                    if description == "":
                        description = None
                    position += 1
                    parent = None
                    if child == "true" and previous_main_criteria:
                        parent = previous_main_criteria
                    info = OptamosCriteria.objects.create(project=project, name=criterion, description=description, position=position, parent=parent)
                    if not info.parent:
                        previous_main_criteria = info

        messages.success(request, "Changes have been saved.")
        return redirect(reverse("optamos:project", args=[project.uid]))

    context = {
        "bg": random.choice(OPTAMOS_BG),
        "projects": project,
        "menu": "projects",
        "project": project,
        "has_descriptions": OptamosCriteria.objects.filter(project=project, description__isnull=False).exists(),
    }
    return render(request, "optamos/project.settings.html", context)

def project_overview(request, id):

    if not request.user.is_authenticated:
        return redirect(f"{reverse("optamos:login")}?redirect={request.path}")

    project = OptamosProject.objects_include_private.get(pk=id)
    user_access = OptamosUser.objects.filter(project=project, user=request.user, level="admin")

    if not user_access.exists():
        return redirect(reverse("optamos:access_denied") + "?role=admin&url=" + request.get_full_path())

    context = {
        "bg": random.choice(OPTAMOS_BG),
        "projects": project,
        "menu": "projects",
        "project": project,
    }
    return render(request, "optamos/project.overview.html", context)

def project_team_results(request, id, page="rank_all_criteria"):

    if not request.user.is_authenticated:
        return redirect(f"{reverse("optamos:login")}?redirect={request.path}")

    project = OptamosProject.objects_include_private.get(pk=id)
    user_access = OptamosUser.objects.filter(project=project, user=request.user, level="admin")
    if not user_access.exists():
        return redirect(reverse("optamos:access_denied") + "?role=admin&url=" + request.get_full_path())

    page = request.GET.get("page", page)
    if "criteria" in request.GET:
        page = "criteria"

    pairs = None
    users = []
    values = {}
    criteria = None

    for each in OptamosUser.objects.filter(project=project).order_by("user__first_name"):
        values[each.user.first_name] = {}
        users.append(each.user.first_name)
    # Also add a new entry for the average
    values["AVERAGE"] = {} 
    users.append("AVERAGE")

    if page == "rank_all_criteria":

        scores = OptamosCriteriaValue.objects.filter(criteria1__project=project)

        for each in scores:
            user = "AVERAGE" if not each.user else each.user.first_name
            label = f"range-{each.criteria1_id}-{each.criteria2_id}"
            values[user][label] = each.value

        # This creates pairs of all possible combinations of alternatives
        pairs = list(combinations(project.criteria.all(), 2))

    elif (criteria := request.GET.get("criteria")):
        page = "criteria"
        criteria = OptamosCriteria.objects.get(project=project, pk=criteria)

        scores = OptamosAlternativeValue.objects.filter(criteria=criteria)

        for each in scores:
            user = "AVERAGE" if not each.user else each.user.first_name
            label = f"range-{each.alternative1_id}-{each.alternative2_id}"
            values[user][label] = each.value

        # This creates pairs of all possible combinations of alternatives
        pairs = list(combinations(project.alternatives.all(), 2))

    context = {
        "bg": random.choice(OPTAMOS_BG),
        "menu": "projects",
        "project": project,
        "criteria_list": project.criteria.all().order_by("id"),
        "remove_padding_main_container": True,
        "page": page,
        "pairs": pairs,
        "values": values,
        "users": users,
        "criteria": criteria,
    }
    return render(request, "optamos/team.results.html", context)

def project_team(request, id):

    if not request.user.is_authenticated:
        return redirect(f"{reverse("optamos:login")}?redirect={request.path}")

    project = OptamosProject.objects_include_private.get(pk=id)
    user_access = OptamosUser.objects.filter(project=project, user=request.user, level="admin")
    if not user_access.exists():
        return redirect(reverse("optamos:access_denied") + "?role=admin&url=" + request.get_full_path())

    if "delete" in request.GET:
        user = User.objects.get(pk=request.GET["delete"])
        people = user.people
        optamos_user = OptamosUser.objects.filter(project=project, user=user)
        if not optamos_user:
            messages.warning(request, f"User was not found within the team and could therefore not be removed.")
        else:
            if people.meta_data and "pending_activation" in people.meta_data:
                # This user was never activated so let's delete this account entirely
                people.delete()
                user.delete()
            else:
                optamos_user.delete()
            messages.success(request, f"The following user was removed from the team: <strong>{user}</strong>")
        return redirect(request.path)

    if request.method == "POST":

        subject = f"Invitation to join project '{project}'"
        sender = '"OPTamos" <' + settings.DEFAULT_FROM_EMAIL + '>'

        emails_raw = request.POST.get("users", "")
        # split on any newline
        raw_lines = re.split(r"\r\n|\r|\n", emails_raw)
        # strip each substring and remove empty lines
        lines = [line.strip() for line in raw_lines]
        lines = [line for line in lines if line]

        # upfront error if too many lines — do not process further
        # We use 26 as the total number of users because it's 1 admin + 25 people max
        remaining = 26 - project.users.count()
        if len(lines) > remaining:
            messages.error(request, f"You can not have more than 25 people in your team. Make sure there is one e-mail per line; no more than {remaining} remaining spots.")
        else:
            invalid = []
            valid = []
            for i, email in enumerate(lines, start=1):
                try:
                    validate_email(email)
                    valid.append(email)
                except ValidationError:
                    messages.error(request, f"The following address is not valid and was skipped: <strong>{email}</strong>")

            for email in valid:
                if (user := User.objects.filter(email=email).first()):
                    # User already has an active account. Check if already part of the project...
                    # And if not, we add them...
                    if project.users.filter(user__pk=user.pk).exists():
                        messages.warning(request, f"The following user was already part of the team: <strong>{email}</strong>")
                    else:
                        messages.success(request, f"The following user has an existing account and was added successfully: <strong>{email}</strong>")
                        OptamosUser.objects.create(project=project, user=user, level="regular")

                        mailcontext = {
                            "project": project,
                            "request": request,
                        }
                        msg_html = render_to_string("mailbody/optamos.existinguser.html", mailcontext)
                        msg_plain = render_to_string("mailbody/optamos.existinguser.txt", mailcontext)
                        send_mail(subject, msg_plain, sender, [email], html_message=msg_html)
                else:
                    password = "".join(random.choices(string.ascii_letters + string.digits, k=20))
                    user = User.objects.create_user(email, email, password)
                    user.first_name = email
                    user.is_superuser = False
                    user.is_staff = False
                    user.save()
                    people = People.objects.create(name=email, email=user.email, user=user, meta_data={"optamos": True, "auto_created": True, "pending_activation": True})
                    OptamosUser.objects.create(project=project, user=user, level="regular")
                    messages.success(request, f"The following user was invited successfully: <strong>{email}</strong>")

                    mailcontext = {
                        "project": project,
                        "request": request,
                        "password": password,
                        "email": email,
                    }
                    msg_html = render_to_string("mailbody/optamos.newuser.html", mailcontext)
                    msg_plain = render_to_string("mailbody/optamos.newuser.txt", mailcontext)
                    send_mail(subject, msg_plain, sender, [email], html_message=msg_html)

            return redirect(request.path)

    context = {
        "bg": random.choice(OPTAMOS_BG),
        "projects": project,
        "menu": "projects",
        "project": project,
        "team": OptamosUser.objects.filter(project=project, level="regular"),
    }
    return render(request, "optamos/project.team.html", context)

def project(request, id, page="home"):

    if not request.user.is_authenticated:
        return redirect(f"{reverse("optamos:login")}?redirect={request.path}")

    if not request.GET:
        return redirect(f"{request.path}?rank_all_criteria=true")

    project = OptamosProject.objects_include_private.filter(pk=id, users__user=request.user).first()
    if not project:
        messages.error(request, "Project is not found - either it does not exist or you do not have access. Below are your projects.")
        return redirect("optamos:projects")

    values = {}
    pairs = None
    sub_pairs = {}
    sub_criteria_pairs = 0

    # This creates a list of pairs grouped by criteria, for all the sub-criteria that are entered
    # We need this for the navigation menu so that's why we do it here
    for each in project.criteria.annotate(child_count=Count("children")).filter(child_count__gt=0):
        sub_combos = list(combinations(project.criteria.filter(parent=each), 2))
        sub_pairs[each.name] = sub_combos
        sub_criteria_pairs += len(sub_combos)

    if (criteria := request.GET.get("criteria")):
        page = "criteria"
        criteria = OptamosCriteria.objects.get(project=project, pk=criteria)

        # If this is a parent that has children, then we will not rank the parent 
        # but only the children, so let's redirect in that case
        if criteria.children.all():
            child = criteria.children.all().order_by("position").first()
            return redirect(reverse("optamos:project", args=[project.uid]) + f"?criteria={child.id}")

        # Let's create a dict with the names of the <input> fields and the value for them
        # so we can load them into the form
        for each in OptamosAlternativeValue.objects.filter(criteria=criteria, user=request.user):
            value = f"range-{each.alternative1_id}-{each.alternative2_id}"
            values[value] = each.js_value

        # This creates pairs of all possible combinations of alternatives
        pairs = list(combinations(project.alternatives.all(), 2))

    elif "rank_all_criteria" in request.GET:
        page = "rank_all_criteria"
        # Let's create a dict with the names of the <input> fields and the value for them
        # so we can load them into the form
        for each in OptamosCriteriaValue.objects.filter(criteria1__project=project, user=request.user):
            value = f"range-{each.criteria1_id}-{each.criteria2_id}"
            values[value] = each.js_value

        # This creates pairs of all possible combinations of criteria
        pairs = list(combinations(project.criteria.filter(parent__isnull=True), 2))

    if request.method == "POST":
        if page == "criteria":
            OptamosAlternativeValue.objects.filter(criteria=criteria, user=request.user).delete()
            for alternative1,alternative2 in pairs:
                # This creates the name of the relevant input field
                value = f"range-{alternative1.id}-{alternative2.id}"
                OptamosAlternativeValue.objects.create(
                    alternative1 = alternative1,
                    alternative2 = alternative2,
                    criteria = criteria,
                    value = ranking_value_converter(int(float(request.POST[value]))),
                    user = request.user,
                )
            if "back" in request.POST:
                next_criteria = project.criteria.filter(position__lt=criteria.position, children__isnull=True).order_by("-position").first()
                if not next_criteria:
                    return redirect(reverse("optamos:project", args=[project.uid]) + f"?rank_all_criteria")
            else:
                next_criteria = project.criteria.filter(position__gt=criteria.position).order_by("position").first()
            if next_criteria:
                return redirect(reverse("optamos:project", args=[project.uid]) + f"?criteria={next_criteria.id}")
            else:
                return redirect(reverse("optamos:project_results", args=[project.uid]))

        elif page == "rank_all_criteria":
            OptamosCriteriaValue.objects.filter(criteria1__project=project, user=request.user).delete()
            for criteria1,criteria2 in pairs:
                # This creates the name of the relevant input field
                value = f"range-{criteria1.id}-{criteria2.id}"
                OptamosCriteriaValue.objects.create(
                    criteria1 = criteria1,
                    criteria2 = criteria2,
                    value = ranking_value_converter(int(float(request.POST[value]))),
                    user = request.user,
                )
            for key, value in sub_pairs.items():
                for criteria1,criteria2 in value:
                    # This creates the name of the relevant input field
                    value = f"range-{criteria1.id}-{criteria2.id}"
                    OptamosCriteriaValue.objects.create(
                        criteria1 = criteria1,
                        criteria2 = criteria2,
                        value = ranking_value_converter(int(float(request.POST[value]))),
                        user = request.user,
                    )
            next_criteria = project.criteria.order_by("position").first()
            if next_criteria:
                return redirect(reverse("optamos:project", args=[project.uid]) + f"?criteria={next_criteria.id}")
            else:
                return redirect(reverse("optamos:project_results", args=[project.uid]))

    context = {
        "bg": random.choice(OPTAMOS_BG),
        "project": project,
        "remove_padding_main_container": True,
        "criteria": criteria,
        "pairs": pairs,
        "sub_pairs": sub_pairs,
        "values": values,
        "page": page,
        "criteria_list": project.criteria.all().annotate(is_done=Count("alternative_pairs", filter=Q(alternative_pairs__user=request.user))).annotate(has_children=Count("children")).order_by("position"),
        "criteria_values": OptamosCriteriaValue.objects.filter(criteria1__project=project, user=request.user).count(), 
        # Count how many there theoretically are, so that we can verify that all are saved -- this is particularly 
        # relevant in case people edit the project and add criteria in which case we need to show an error
        "total_required_criteria_values": len(list(combinations(project.criteria.filter(parent__isnull=True), 2))) + sub_criteria_pairs,
        "menu": "projects",
        "cr": calculate_consistency_ratio(list(project.criteria.all()), OptamosCriteriaValue.objects.filter(criteria1__project=project, user=request.user)).cr,
        "next_criteria": project.criteria.filter(pk__gt=criteria.pk).order_by("id").first() if criteria else None,
        "criteria_descriptions": OptamosCriteria.objects.filter(project=project, description__isnull=False),
        "access_level": OptamosUser.objects.get(user=request.user, project=project).level,
    }
    return render(request, "optamos/project.html", context)

def project_results(request, id, page="results", team=False):

    if not request.user.is_authenticated:
        return redirect(f"{reverse("optamos:login")}?redirect={request.path}")

    project = OptamosProject.objects_include_private.filter(pk=id, users__user=request.user).first()
    if not project:
        messages.error(request, "Project is not found - either it does not exist or you do not have access. Below are your projects.")
        return redirect("optamos:projects")

    user = None if team else request.user

    # This creates a list of pairs grouped by criteria, for all the sub-criteria that are entered
    # We need this for the navigation menu so that's why we do it here
    sub_criteria_pairs = 0
    for each in project.criteria.annotate(child_count=Count("children")).filter(child_count__gt=0):
        sub_combos = list(combinations(project.criteria.filter(parent=each), 2))
        sub_criteria_pairs += len(sub_combos)

    # ---------------------------------------------
    # START OF HIERARCHICAL AHP CALCULATION
    # ---------------------------------------------

    # Get all criteria and alternatives
    criteria = list(OptamosCriteria.objects.filter(project=project))
    alternatives = list(OptamosAlternative.objects.filter(project=project))

    # ---------------------------------------------
    # STEP 1: BUILD HIERARCHY STRUCTURE
    # ---------------------------------------------
    children_map = {}
    for c in criteria:
        parent_id = c.parent.id if c.parent else None
        children_map.setdefault(parent_id, []).append(c)

    root_criteria = children_map.get(None, [])
    leaf_criteria = [c for c in criteria if c.id not in children_map]

    # ---------------------------------------------
    # STEP 2: BUILD MATRIX FOR CRITERIA SUBSET
    # ---------------------------------------------
    def create_matrix_for_subset(criteria_subset):
        matrix = {
            c1.id: {c2.id: 1 if c1.id == c2.id else 0 for c2 in criteria_subset}
            for c1 in criteria_subset
        }

        values = OptamosCriteriaValue.objects.filter(
            criteria1__in=criteria_subset,
            criteria2__in=criteria_subset,
            user=user
        )

        for v in values:
            c1 = v.criteria1.id
            c2 = v.criteria2.id
            matrix[c1][c2] = v.value1
            matrix[c2][c1] = v.value2

        return matrix

    # ---------------------------------------------
    # STEP 3: COMPUTE LOCAL WEIGHTS PER NODE
    # ---------------------------------------------
    criteria_weights_local = {}

    def compute_local_weights(criteria_subset):
        if len(criteria_subset) == 1:
            return {criteria_subset[0].id: 1.0}

        matrix = create_matrix_for_subset(criteria_subset)

        col_totals = {c.id: 0 for c in criteria_subset}
        for col in criteria_subset:
            for row in criteria_subset:
                col_totals[col.id] += matrix[row.id][col.id]

        normalized = {}
        for row in criteria_subset:
            normalized[row.id] = {}
            for col in criteria_subset:
                total = col_totals[col.id]
                normalized[row.id][col.id] = matrix[row.id][col.id] / total if total else 0

        weights = {}
        for row in criteria_subset:
            weights[row.id] = sum(normalized[row.id][col.id] for col in criteria_subset) / len(criteria_subset)

        return weights

    def compute_all_local_weights(parent_id=None):
        siblings = children_map.get(parent_id, [])
        if not siblings:
            return

        local_weights = compute_local_weights(siblings)

        for c in siblings:
            criteria_weights_local[c.id] = local_weights[c.id]
            compute_all_local_weights(c.id)

    compute_all_local_weights()

    # ---------------------------------------------
    # STEP 4: COMPUTE GLOBAL WEIGHTS
    # ---------------------------------------------
    criteria_weights_global = {}

    def compute_global_weights(parent_id=None, parent_weight=1.0):
        siblings = children_map.get(parent_id, [])
        for c in siblings:
            local_w = criteria_weights_local.get(c.id, 1)
            global_w = parent_weight * local_w
            criteria_weights_global[c.id] = global_w

            compute_global_weights(c.id, global_w)

    compute_global_weights()

    # ---------------------------------------------
    # STEP 5: ALTERNATIVE EVALUATION
    # ---------------------------------------------
    alternative_values = OptamosAlternativeValue.objects.filter(criteria__project=project, user=user)

    alternative_matrices = {}
    for c in criteria:
        alternative_matrices[c.id] = {
            o1.id: {o2.id: 1 if o1.id == o2.id else 0 for o2 in alternatives}
            for o1 in alternatives
        }

    for ov in alternative_values:
        c_id = ov.criteria.id
        o1_id = ov.alternative1.id
        o2_id = ov.alternative2.id

        alternative_matrices[c_id][o1_id][o2_id] = ov.value1
        alternative_matrices[c_id][o2_id][o1_id] = ov.value2

    normalized_alternative_matrices = {}
    alternative_weights = {}

    for c in criteria:
        matrix = alternative_matrices[c.id]

        col_totals = {o.id: 0 for o in alternatives}
        for col in alternatives:
            for row in alternatives:
                col_totals[col.id] += matrix[row.id][col.id]

        normalized_matrix = {}
        for row in alternatives:
            normalized_matrix[row.id] = {}
            for col in alternatives:
                total = col_totals[col.id]
                normalized_matrix[row.id][col.id] = matrix[row.id][col.id] / total if total else 0

        normalized_alternative_matrices[c.id] = normalized_matrix

        row_avg = {}
        for row in alternatives:
            total = sum(normalized_matrix[row.id][col.id] for col in alternatives)
            row_avg[row.id] = total / len(alternatives)

        alternative_weights[c.id] = row_avg

    # ---------------------------------------------
    # STEP 6: GLOBAL SCORING (LEAF ONLY!)
    # ---------------------------------------------
    global_scores = {o.id: 0 for o in alternatives}

    for o in alternatives:
        total_score = 0
        for c in leaf_criteria:
            weight_c = criteria_weights_global[c.id]
            weight_o = alternative_weights[c.id][o.id]
            total_score += weight_c * weight_o

        global_scores[o.id] = total_score

    global_ranking = sorted(
        [{'alternative': o, 'score': global_scores[o.id]} for o in alternatives],
        key=lambda x: x['score'],
        reverse=True
    )

    # ---------------------------------------------
    # STEP 7: CONSISTENCY RATIOS (UNCHANGED)
    # ---------------------------------------------
    consistency = calculate_consistency_ratio(criteria, OptamosCriteriaValue.objects.filter(criteria1__project=project, user=user))

    alternative_crs = {}
    for c in criteria:
        alternative_crs[c.id] = calculate_consistency_ratio(alternatives, alternative_values.filter(criteria=c)).cr

    # ---------------------------------------------
    # STEP 8: SUMMARY TABLE (LEAF ONLY!)
    # ---------------------------------------------
    grouped_summary = []
    importance = []

    # Separate standalone and grouped criteria
    groups = {}

    for c in leaf_criteria:
        if c.parent:
            groups.setdefault(c.parent, []).append(c)
        else:
            groups.setdefault(c, []).append(c)  # acts as its own group

    for parent, children in groups.items():
        group = {
            "parent": parent.name,
            "is_group": any(child.parent for child in children),  # True if real parent
            "rows": []
        }

        for c in children:
            row = {
                "child": c.name,
                "child_id": c.id,
                "alternatives": [
                    alternative_weights[c.id][o.id] * criteria_weights_global[c.id] * 100
                    for o in alternatives
                ],
                "cr": alternative_crs[c.id],
                "importance": criteria_weights_global[c.id] * 100
            }

            importance.append({
                "id": c.id,
                "name": f"{parent.name} → {c.name}" if c.parent else c.name,
                "value": row["importance"]
            })

            group["rows"].append(row)

        grouped_summary.append(group)

    # Compute totals
    totals = []
    for idx in range(len(alternatives)):
        total = sum(
            row["alternatives"][idx]
            for group in grouped_summary
            for row in group["rows"]
        )
        totals.append(total)

    # ---------------------------------------------
    # END OF HIERARCHICAL AHP CALCULATION
    # ---------------------------------------------

    if "export" in request.GET:

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Summary Table"

        bold_font = Font(bold=True)

        # ----------------------------------------
        # HEADER
        # ----------------------------------------
        header = ["Criterion"] + [o.name for o in alternatives] + ["CR", "% Importance"]
        worksheet.append(header)

        for cell in worksheet[1]:
            cell.font = bold_font

        # ----------------------------------------
        # DATA
        # ----------------------------------------
        for group in grouped_summary:
            parent_name = group["parent"]
            is_group = group["is_group"]
            rows = group["rows"]

            if is_group:
                # Parent row (sum of children)
                parent_alternatives = [
                    sum(row["alternatives"][i] for row in rows)
                    for i in range(len(alternatives))
                ]
                parent_importance = sum(row["importance"] for row in rows)

                worksheet.append(
                    [parent_name] +
                    [val / 100 for val in parent_alternatives] +
                    ["", parent_importance / 100]
                )

                # Children rows
                for row in rows:
                    worksheet.append(
                        [f"  → {row['child']}"] +
                        [val / 100 for val in row["alternatives"]] +
                        [row["cr"], row["importance"] / 100]
                    )

            else:
                # Single (no hierarchy)
                row = rows[0]
                worksheet.append(
                    [parent_name] +
                    [val / 100 for val in row["alternatives"]] +
                    [row["cr"], row["importance"] / 100]
                )

        # ----------------------------------------
        # TOTALS
        # ----------------------------------------
        worksheet.append(
            ["Totals"] +
            [val / 100 for val in totals] +
            ["", ""]
        )

        # ----------------------------------------
        # FORMATTING
        # ----------------------------------------
        for row in worksheet.iter_rows(min_row=2, min_col=2):
            for i, cell in enumerate(row):
                if isinstance(cell.value, (int, float)):
                    # Alternatives + importance → %
                    if i < len(alternatives) or i == len(alternatives) + 1:
                        cell.number_format = "0.0%"
                    else:  # CR column
                        cell.number_format = "0.00"

        # Bold first column
        for cell in worksheet["A"]:
            cell.font = bold_font

        # Bold totals row
        for cell in worksheet[worksheet.max_row]:
            cell.font = bold_font

        # ----------------------------------------
        # RESPONSE
        # ----------------------------------------
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        filename = f"Summary table - {project.name}.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        return response

    elif "export_full" in request.GET:

        workbook = openpyxl.Workbook()
        bold = Font(bold=True)

        def auto_width(ws):
            for col in ws.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = max_length + 2


        # --------------------------------------------------
        # 1. SUMMARY TABLE
        # --------------------------------------------------
        ws = workbook.active
        ws.title = "Summary"

        ws.append(["Criterion"] + [o.name for o in alternatives] + ["CR", "Importance (%)"])
        for cell in ws[1]:
            cell.font = bold

        current_row = 2

        for group in grouped_summary:
            parent_name = group["parent"]
            is_group = group["is_group"]
            rows = group["rows"]

            if is_group:
                parent_alternatives = [
                    sum(row["alternatives"][i] for row in rows)
                    for i in range(len(alternatives))
                ]
                parent_importance = sum(row["importance"] for row in rows)

                ws.append(
                    [parent_name] +
                    [v / 100 for v in parent_alternatives] +
                    ["", parent_importance / 100]
                )

                parent_excel_row = current_row
                current_row += 1

                for row in rows:
                    ws.append(
                        [f"  → {row['child']}"] +
                        [v / 100 for v in row["alternatives"]] +
                        [row["cr"], row["importance"] / 100]
                    )
                    ws.row_dimensions[current_row].outlineLevel = 1
                    ws.row_dimensions[current_row].hidden = True
                    current_row += 1

                ws.row_dimensions[parent_excel_row].collapsed = True

            else:
                row = rows[0]
                ws.append(
                    [parent_name] +
                    [v / 100 for v in row["alternatives"]] +
                    [row["cr"], row["importance"] / 100]
                )
                current_row += 1

        # totals
        ws.append(["Totals"] + [v / 100 for v in totals] + ["", ""])
        for cell in ws[ws.max_row]:
            cell.font = bold

        # formatting
        for row in ws.iter_rows(min_row=2, min_col=2):
            for i, cell in enumerate(row):
                if isinstance(cell.value, (int, float)):
                    if i < len(alternatives) or i == len(alternatives) + 1:
                        cell.number_format = "0.0%"
                    else:
                        cell.number_format = "0.00"

        ws.freeze_panes = "A2"
        auto_width(ws)

        # --------------------------------------------------
        # 2. GLOBAL RANKING
        # --------------------------------------------------
        ws = workbook.create_sheet("Global Ranking")

        ws.append(["Rank", "Alternative", "Score (%)"])
        for cell in ws[1]:
            cell.font = bold

        for idx, item in enumerate(global_ranking, start=1):
            ws.append([idx, item["alternative"].name, item["score"]])

        for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
            for cell in row:
                cell.number_format = "0.0%"

        ws.freeze_panes = "A2"
        auto_width(ws)


        # --------------------------------------------------
        # 3. CRITERIA WEIGHTS 
        # --------------------------------------------------
        ws = workbook.create_sheet("Criteria Weights")

        ws.append(["Criterion", "Weight", "Percentage"])
        for cell in ws[1]:
            cell.font = bold

        current_row = 2

        for group in grouped_summary:
            parent = group["parent"]
            is_group = group["is_group"]
            rows = group["rows"]

            if is_group:
                parent_weight = sum(row["importance"] for row in rows)

                ws.append([parent, parent_weight / 100, parent_weight / 100])
                parent_excel_row = current_row
                current_row += 1

                for row in rows:
                    ws.append([
                        f"  → {row['child']}",
                        row["importance"] / 100,
                        row["importance"] / 100
                    ])
                    ws.row_dimensions[current_row].outlineLevel = 1
                    ws.row_dimensions[current_row].hidden = True
                    current_row += 1

                ws.row_dimensions[parent_excel_row].collapsed = True

            else:
                row = rows[0]
                ws.append([
                    parent,
                    row["importance"] / 100,
                    row["importance"] / 100
                ])
                current_row += 1

        for row in ws.iter_rows(min_row=2, min_col=2):
            for i, cell in enumerate(row):
                if i == 0:
                    cell.number_format = "0.000"
                else:
                    cell.number_format = "0.0%"

        ws.freeze_panes = "A2"
        auto_width(ws)


        # --------------------------------------------------
        # 4–6. ALTERNATIVE MATRICES
        # --------------------------------------------------
        for c in criteria:
            # RAW
            ws = workbook.create_sheet(f"{c.name} – Raw")
            ws.append([""] + [o.name for o in alternatives])
            for cell in ws[1]:
                cell.font = bold

            for r in alternatives:
                ws.append(
                    [r.name] +
                    [alternative_matrices[c.id][r.id][col.id] for col in alternatives]
                )

            ws.freeze_panes = "B2"
            auto_width(ws)

            # NORMALIZED
            ws = workbook.create_sheet(f"{c.name} – Normalized")
            ws.append([""] + [o.name for o in alternatives])
            for cell in ws[1]:
                cell.font = bold

            for r in alternatives:
                ws.append(
                    [r.name] +
                    [
                        normalized_alternative_matrices[c.id][r.id][col.id]
                        for col in alternatives
                    ]
                )

            ws.freeze_panes = "B2"
            auto_width(ws)

            # WEIGHTS
            ws = workbook.create_sheet(f"{c.name} – Weights")
            ws.append(["Alternative", "Weight", "Percentage"])
            for cell in ws[1]:
                cell.font = bold

            for o in alternatives:
                w = alternative_weights[c.id][o.id]
                ws.append([o.name, w, w])

            for row in ws.iter_rows(min_row=2, min_col=2):
                for i, cell in enumerate(row):
                    if i == 0:
                        cell.number_format = "0.000"
                    else:
                        cell.number_format = "0.0%"

            ws.freeze_panes = "A2"
            auto_width(ws)


        # --------------------------------------------------
        # RESPONSE
        # --------------------------------------------------
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="Full AHP Results - {project.name}.xlsx"'

        return response

    if request.method == "POST" and page == "sensitivity":

        # ---------------------------------------------
        # SENSITIVITY VIEW (HIERARCHY-AWARE)
        # ---------------------------------------------

        crit_id = int(request.POST.get("criterion_id"))
        new_weight = float(request.POST.get("new_weight")) / 100

        # ONLY leaf criteria matter
        adjusted_weights = {
            c.id: criteria_weights_global[c.id]
            for c in leaf_criteria
        }

        old_weight = adjusted_weights[crit_id]
        remaining_weight = 1.0 - new_weight
        old_remaining_weight = 1.0 - old_weight

        # redistribute
        for c_id in adjusted_weights:
            if c_id == crit_id:
                adjusted_weights[c_id] = new_weight
            else:
                adjusted_weights[c_id] = (
                    adjusted_weights[c_id] * remaining_weight / old_remaining_weight
                    if old_remaining_weight > 0 else 0
                )

        # recompute scores (LEAF ONLY!)
        sensitivity_global_scores = {o.id: 0 for o in alternatives}

        for o in alternatives:
            total_score = 0
            for c in leaf_criteria:
                weight_c = adjusted_weights[c.id]
                weight_o = alternative_weights[c.id][o.id]
                total_score += weight_c * weight_o

            sensitivity_global_scores[o.id] = total_score

        sensitivity_ranking = sorted(
            [
                {
                    "alternative_id": o.id,
                    "alternative_name": o.name,
                    "score": sensitivity_global_scores[o.id] * 100
                }
                for o in alternatives
            ],
            key=lambda x: x["score"],
            reverse=True
        )

        return JsonResponse({
            "sensitivity_ranking": sensitivity_ranking,
            "adjusted_weights": {k: v * 100 for k, v in adjusted_weights.items()}
        })


    context = {
        "bg": random.choice(OPTAMOS_BG),
        "project": project,
        "remove_padding_main_container": True,
        "criteria": criteria,
        "page": page,
        "criteria_list": project.criteria.all().annotate(is_done=Count("alternative_pairs", filter=Q(alternative_pairs__user=request.user))).annotate(has_children=Count("children")).order_by("position"),
        "criteria_values": OptamosCriteriaValue.objects.filter(criteria1__project=project, user=request.user).count(), 
        # Count how many there theoretically are, so that we can verify that all are saved -- this is particularly 
        # relevant in case people edit the project and add criteria in which case we need to show an error
        "total_required_criteria_values": len(list(combinations(project.criteria.filter(parent__isnull=True), 2))) + sub_criteria_pairs, 
        "menu": "projects",

        "criteria": criteria,
        "alternatives": alternatives,
        "alternative_matrices": alternative_matrices,
        "normalized_alternative_matrices": normalized_alternative_matrices,
        "alternative_weights": alternative_weights,

        "global_scores": global_scores,
        "global_ranking": global_ranking,
        "alternative_crs": alternative_crs,
        "grouped_summary": grouped_summary,
        "summary_totals": totals,
        "totals": totals,
        "lambda_max": consistency.lambda_max,
        "ci": consistency.ci,
        "cr": consistency.cr,
        "importance": importance,
        "team": team,
        "access_level": OptamosUser.objects.get(user=request.user, project=project).level,
        "project_has_subcriteria": OptamosCriteria.objects.filter(project=project, parent__isnull=False).exists(),
    }

    return render(request, "optamos/project.html", context)

# ACCOUNT-RELATED FUNCTIONS

def account_login(request):

    if request.method == "POST":
        email = request.POST.get("email").lower()

        password = request.POST.get("password")
        user = authenticate(request, username=email.strip(), password=password.strip())
        redirect_url = request.GET.get("redirect", "optamos:projects")

        if user is not None:
            login(request, user)
            people = People.objects.get(user=user)
            if people.meta_data and "temporary_password" in people.meta_data:
                messages.success(request, "Please change your temporary pin. You can set your own password here:" + "<br><a href='/hub/profile/edit/?shortened=true'>" + "Edit my profile" + "</a>")
            elif people.meta_data and "pending_activation" in people.meta_data:
                messages.success(request, "Welcome to OPTamos! Please finish setting up your account here:" + "<br><a href='/account/?activation=true'>" + "Edit my account" + "</a>")
            return redirect(redirect_url)
        else:
            messages.error(request, "We could not authenticate you, please try again.")

    context = {
        "bg": random.choice(OPTAMOS_BG),
        "menu": "login",
    }
    return render(request, "optamos/login.html", context)

def account_logout(request):
    logout(request)
    messages.success(request, "You are now logged out")
    return redirect("optamos:index")

def account(request):
    if not request.user.is_authenticated:
        return redirect(f"{reverse("optamos:login")}?redirect={request.path}")

    if request.method == "POST":
        user = request.user
        people = user.people
        name = request.POST.get("name")
        email = request.POST.get("email")

        if email != user.email and User.objects.filter(email = email).exists():
            messages.error(request, "E-mail already in use; cannot change this e-mail address")
            return redirect(request.path)

        people.name = name
        user.first_name = name
        people.email = email
        user.username = email
        user.email = email

        if "password" in request.POST and request.POST["password"]:
            user.set_password(request.POST["password"])
        user.save();

        if people.meta_data and "pending_activation" in people.meta_data:
            del(people.meta_data["pending_activation"])

        if not people.meta_data:
            people.meta_data = {}

        if "institution" in request.POST:
            people.meta_data["institution"] = request.POST.get("institution")
        if "location" in request.POST:
            people.meta_data["location"] = request.POST.get("location")
        if "how" in request.POST:
            people.meta_data["how"] = request.POST.get("how")
        people.save()

        login(request, user)
        messages.success(request, "Changes have been saved.")
        return redirect(request.path)

    context = {
        "bg": random.choice(OPTAMOS_BG),
        "menu": "account",
    }
    return render(request, "optamos/account.settings.html", context)

def account_create(request):

    redirect_url = request.GET.get("next") if request.GET.get("next") else reverse("optamos:projects")

    if request.user.is_authenticated:
        is_logged_in = True
        return redirect(redirect_url)

    if request.method == "POST":
        error = None
        password = request.POST.get("password")
        email = request.POST.get("email")
        name = request.POST.get("name")
        if request.POST.get("phone").lower() != "good morning":
            messages.error(request, "Please enter 'Good morning' in the last box.")
            error = True
        elif not password:
            messages.error(request, "You did not enter a password.")
            error = True
        elif User.objects.filter(email=email).exists():
            url = reverse("optamos:login")
            messages.error(request, f"An account already exists with this e-mail address. Please <a href='{url}'>log in</a>. <br>Remember that Metabolism of Islands accounts also work to log into OPTamos.")
            error = True

        if not error:
            user = User.objects.create_user(email, email, password)
            user.first_name = name
            user.is_superuser = False
            user.is_staff = False
            user.save()
            login(request, user)

            people = People.objects.create(name=name, email=user.email)
            people.user = user

            meta_data = {"optamos": True}
            if "institution" in request.POST:
                meta_data["institution"] = request.POST.get("institution")
            if "location" in request.POST:
                meta_data["location"] = request.POST.get("location")
            if "how" in request.POST:
                meta_data["how"] = request.POST.get("how")
            people.meta_data = meta_data
            people.save()

            messages.success(request, "You are successfully registered.")

            return redirect(redirect_url)

    context = {
        "bg": random.choice(OPTAMOS_BG),
    }
    return render(request, "optamos/account.html", context)

def access_denied(request):

    context = {
        "bg": random.choice(OPTAMOS_BG),
    }
    return render(request, "optamos/accessdenied.html", context)
